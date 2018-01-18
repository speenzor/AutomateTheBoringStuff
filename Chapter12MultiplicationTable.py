#! usr/bin/env python3
#multiplicationTable.py - Multiplication Table Marker
#Takes the second argument of input, an integer, and makes a multiplication
#table of that size in Excel.

import os, sys, openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
#tableSize = sys.argv[1]
tableSize = 6
fontObj = Font(bold=True)

#Loop to add the top row of 1 to tableSize then left column of save size, in bold
for topCell in range(1,tableSize+1):
    sheet.cell(row=1, column=topCell+1).value = topCell
    sheet.cell(row=1, column=topCell+1).font = fontObj
    sheet.cell(row=topCell+1, column=1).value = topCell
    sheet.cell(row=topCell+1, column=1).font = fontObj

#Loop through and multiply
for x in range(2,tableSize+2):
    for y in range(2,tableSize+2):
        sheet.cell(row=x, column=y).value = sheet.cell(row=x, column=1).value * sheet.cell(row=1, column=y).value

#Save the new workbook
wb.save('multiTable.xlsx')

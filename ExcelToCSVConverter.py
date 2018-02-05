#Automate the Boring Stuff Chapter 14 Challenge- Excel-To-CSV Converter
#This program reads all the Excel files in the current working directory and
#outputs them as CSV files.

import openpyxl, os, csv

path = '/Users/spencercorwin/Desktop/Folder2'
os.chdir('/Users/spencercorwin/Desktop/Folder2')

for excelFile in os.listdir(path):
    #Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue  #skip non-xlsx files
    wb = openpyxl.load_workbook(path+'/'+excelFile)
    for sheetName in wb.get_sheet_names():
        #Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)
        #Create the CSV filename from the Excel filename and sheet title.
        csvFileObj = open(os.path.splitext(excelFile)[0]+'_'+sheetName+'.csv', 'w', newline='')
        #Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFileObj)

        #Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row+1):
            rowData = [] #append each cell to this list
            #Loop through each cell in the row.
            for colNum in range(1, sheet.max_column+1):
                #Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            #Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        #Close the CSV file object.
        csvFileObj.close()
